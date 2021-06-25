import pandas as pd
import openpyxl
def get_result(path):
    df = pd.read_excel(path)
    #df.to_csv(path, index = None, header=True)
    results = []
    df = pd.read_excel(path)
    df = df.to_numpy()

    print(df)
    results = []
    for i, element in enumerate(df):
        x, y, z = element[0], element[1], element[2]
        if z.startswith("add"):
            result = x + y
        if z.startswith("mul"):
            result = x * y
        if z.startswith("exp"):
            result = x ** y
        if z.startswith("sub"):
            result = x - y
        if z.startswith("div"):
            if y != 0:
                result = x / y
            else:
                result = "Error"
        if result == -0.0:
            result = 0.0
        results.append(result)
    return results


def write(path,arr):
    file = openpyxl.load_workbook(path)
    sheet = file.active
    for i,item in enumerate(arr):
        if type(item) is complex:
            sheet.cell(row=i + 2, column=4).value = str(item)
        else:
            sheet.cell(row= i+2,column=4).value = item
    file.save(path)


print("\n")
path = input("Enter xlsx file path: ")
try:
    arr = get_result(path)
    write(path,arr)
except Exception as e:
    print(e)
