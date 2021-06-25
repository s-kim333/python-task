
import openpyxl

print("\n")
path = input("Enter xlsx file path: ")

try:
    wb_obj = openpyxl.load_workbook(path.strip())
    sheet_obj = wb_obj.active

    # get max column count
    max_column=sheet_obj.max_column
    max_row=sheet_obj.max_row

    results = []
    for j in range(2,max_row+1):
        result_cell = sheet_obj.cell(row = j, column =4)
        x_cel = sheet_obj.cell(row=j, column=1)
        y_cel = sheet_obj.cell(row=j, column=2)
        z_cel = sheet_obj.cell(row=j, column=3)

        result = "none"
        if z_cel.value.startswith("add"):
            result  = x_cel.value + y_cel.value
        if z_cel.value.startswith("mul"):
            result  = x_cel.value * y_cel.value
        if z_cel.value.startswith("exp"):
            result  = x_cel.value ** y_cel.value
        if z_cel.value.startswith("sub"):
            result  = x_cel.value - y_cel.value
        if z_cel.value.startswith("div"):
            if y_cel.value != 0:
                result = x_cel.value / y_cel.value
            else:
                result = "Error"
        # handling complex number
        if type(result) == complex:
            result_cell.value = str(result)
        else:
            result_cell.value = result
    # updating file
    wb_obj.save(path)
except Exception as e:
    print(e)
